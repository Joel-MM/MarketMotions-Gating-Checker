import os
import re
import time
from tempfile import NamedTemporaryFile

import streamlit as st
from dotenv import load_dotenv
from openpyxl import load_workbook
from sp_api.api import ListingsRestrictions
from sp_api.base import Marketplaces, SellingApiException

load_dotenv()

SELLER_ID = os.getenv("SELLER_ID")
MARKETPLACE_ID = os.getenv("MARKETPLACE_ID", "A2EUQ1WTGCTBG2")

credentials = {
    "refresh_token": os.getenv("REFRESH_TOKEN"),
    "lwa_app_id": os.getenv("LWA_APP_ID"),
    "lwa_client_secret": os.getenv("LWA_CLIENT_SECRET"),
    "aws_access_key": os.getenv("AWS_ACCESS_KEY_ID"),
    "aws_secret_key": os.getenv("AWS_SECRET_ACCESS_KEY"),
}

api = ListingsRestrictions(
    marketplace=Marketplaces.CA,
    credentials=credentials
)

def extract_asin(value):
    text = str(value or "").strip()

    patterns = [
        r"/dp/([A-Z0-9]{10})",
        r"/gp/product/([A-Z0-9]{10})",
        r"/product/([A-Z0-9]{10})",
        r"asin=([A-Z0-9]{10})",
        r"^([A-Z0-9]{10})$"
    ]

    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return match.group(1).upper()

    return None

def check_gated(asin):
    response = api.get_listings_restrictions(
        asin=asin,
        sellerId=SELLER_ID,
        marketplaceIds=[MARKETPLACE_ID],
        conditionType="new_new"
    )

    restrictions = response.payload.get("restrictions", [])
    return "Yes" if restrictions else "No"

def process_keepa_file(input_path, output_path):
    workbook = load_workbook(input_path)
    sheet = workbook.active

    sheet.insert_cols(6)
    sheet["F1"] = "Gated?"

    cache = {}
    total_rows = sheet.max_row - 1
    progress_bar = st.progress(0)
    status_text = st.empty()

    checked_count = 0

    for row in range(2, sheet.max_row + 1):
        amazon_url = sheet[f"E{row}"].value
        asin = extract_asin(amazon_url)

        checked_count += 1
        progress_bar.progress(checked_count / total_rows)
        status_text.write(f"Checking row {row} of {sheet.max_row}...")

        if not asin:
            sheet[f"F{row}"] = "No ASIN found"
            continue

        if asin in cache:
            sheet[f"F{row}"] = cache[asin]
            continue

        try:
            result = check_gated(asin)
            cache[asin] = result
            sheet[f"F{row}"] = result
            time.sleep(0.25)

        except SellingApiException:
            sheet[f"F{row}"] = "API Error"

        except Exception:
            sheet[f"F{row}"] = "Error"

    workbook.save(output_path)
    status_text.write("Done.")
    return output_path

st.set_page_config(
    page_title="MarketMotions Gating Checker",
    page_icon="📦",
    layout="centered"
)

st.title("MarketMotions Gating Checker")
st.write("Upload a Keepa Excel export. The app will check Amazon.ca listing restrictions and write Yes/No in column F.")

uploaded_file = st.file_uploader("Upload Keepa Excel file", type=["xlsx"])

if uploaded_file:
    st.success("File uploaded.")

    if st.button("Check Gating"):
        with NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_input:
            temp_input.write(uploaded_file.read())
            input_path = temp_input.name

        output_path = input_path.replace(".xlsx", "_checked.xlsx")

        try:
            process_keepa_file(input_path, output_path)

            with open(output_path, "rb") as file:
                st.download_button(
                    label="Download checked file",
                    data=file,
                    file_name="Keepa_checked.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

        except Exception as error:
            st.error(f"Something went wrong: {error}")