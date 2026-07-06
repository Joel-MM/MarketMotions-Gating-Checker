import os
import re
import time
from dotenv import load_dotenv
from openpyxl import load_workbook
from sp_api.api import ListingsRestrictions
from sp_api.base import Marketplaces, SellingApiException

load_dotenv()

SELLER_ID = os.getenv("SELLER_ID")
MARKETPLACE_ID = os.getenv("MARKETPLACE_ID", "A2EUQ1WTGCTBG2")

credentials = {
    "refresh_token": os.getenv("REFRESH_TOKEN"),
    "lwa_app_id": os.getenv("LWA_APP_ID"),
    "lwa_client_secret": os.getenv("LWA_CLIENT_SECRET"),
    "aws_access_key": os.getenv("AWS_ACCESS_KEY_ID"),
    "aws_secret_key": os.getenv("AWS_SECRET_ACCESS_KEY"),
}

api = ListingsRestrictions(
    marketplace=Marketplaces.CA,
    credentials=credentials
)

def extract_asin(value):
    text = str(value or "").strip()

    patterns = [
        r"/dp/([A-Z0-9]{10})",
        r"/gp/product/([A-Z0-9]{10})",
        r"/product/([A-Z0-9]{10})",
        r"asin=([A-Z0-9]{10})",
        r"^([A-Z0-9]{10})$"
    ]

    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return match.group(1).upper()

    return None

def check_gated(asin):
    response = api.get_listings_restrictions(
        asin=asin,
        sellerId=SELLER_ID,
        marketplaceIds=[MARKETPLACE_ID],
        conditionType="new_new"
    )

    restrictions = response.payload.get("restrictions", [])
    return "Yes" if restrictions else "No"

def process_file(input_file):
    workbook = load_workbook(input_file)
    sheet = workbook.active

    sheet.insert_cols(6)
    sheet["F1"] = "Gated?"

    cache = {}

    for row in range(2, sheet.max_row + 1):
        amazon_url = sheet[f"E{row}"].value
        asin = extract_asin(amazon_url)

        if not asin:
            sheet[f"F{row}"] = "No ASIN found"
            continue

        if asin in cache:
            sheet[f"F{row}"] = cache[asin]
            continue

        try:
            result = check_gated(asin)
            cache[asin] = result
            sheet[f"F{row}"] = result
            print(f"Row {row}: {asin} = {result}")

            time.sleep(0.25)

        except SellingApiException as error:
            sheet[f"F{row}"] = "API Error"
            print(f"Row {row}: {asin} = API Error: {error}")

        except Exception as error:
            sheet[f"F{row}"] = "Error"
            print(f"Row {row}: {asin} = Error: {error}")

    output_file = input_file.replace(".xlsx", "_checked.xlsx")
    workbook.save(output_file)

    print("\nDone.")
    print(f"Saved checked file as: {output_file}")

if __name__ == "__main__":
    file_path = input("Enter the Keepa Excel file path: ").strip().strip('"')
    process_file(file_path)